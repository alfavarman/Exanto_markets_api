import os
import sys
import logging
from api_services import get_api_data, get_api_data_mock
from excel import ExelHandler
from openpyxl import load_workbook

# set up logging
logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO, datefmt="%Y-%m-%d %H:%M:%S")


def main():
    # if file_path name is not given use default file_path
    file_path = sys.argv[1] if len(sys.argv) == 2 else "portfel.xlsx"
    # check if file exists
    if not os.path.isfile(file_path):
        sys.exit(f"File {file_path} not found")

    # get API data
    try:
        api_data = get_api_data()
    except Exception as e:
        logging.error(
            f"Connecting to API FAILED!"
            f"\n\t{e}"
            f"\n\tHave You updated .env with credentials?!"
            f"\n\tYou will use outdated data!"
        )
        api_data = get_api_data_mock()  # Monkey patch in case of API failure
        logging.info(f"You are using test-data (from data.json file)")

    # with context manager to handle errors
    with ExelHandler(file_name=file_path, load_workbook_func=load_workbook) as file_handler:
        file_handler.update_file_with_api_data(api_data=api_data)


if __name__ == "__main__":
    main()
