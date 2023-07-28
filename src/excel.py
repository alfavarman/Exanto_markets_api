import logging
from typing import List, Dict
from openpyxl import load_workbook


class ExelHandler:
    def __init__(self, file_name: str):
        self.file_name = file_name
        self.EXCLUDED_TICKERS = ("BTC", "ETH", "USD", "TLT", "BUSD")
        self.AVAILABLE_CASH = "Dolar amerykański (gotówka) Exante"
        self.available_cash_row_col = None
        # ticker column
        self.COL_B = 2
        # exposition column
        self.COL_E = 5
        self.sheet = None
        self.positions = None

    def load_file_data(self) -> None:
        self._load_available_cash_row()
        self._load_file_positions()
        self._load_active_sheet(file_name=self.file_name)

    def update_file_with_api_data(self, api_data: dict) -> None:
        """ update position to file"""
        api_positions = api_data["positions"]

        for position in api_positions:
            self._update_position(api_position=position)

        self._update_file_exante_available_cash(api_data["currencies"])

    def save_and_close(self):
        self.sheet.save()
        self.sheet.close()
        logging.info(f"File {self.file_name} updated successfully")

    def _load_available_cash_row(self) -> None:
        for row_number, row in enumerate(self.sheet.iter_rows(values_only=True), start=1):
            if row[0] == self.AVAILABLE_CASH:
                self.available_cash_row = row_number
                break
        else:
            raise Exception(f"{self.AVAILABLE_CASH} not found in file")

    def _load_active_sheet(self, file_name: str) -> None:
        wb = load_workbook(file_name)
        self.sheet = wb.active

    def _load_file_positions(self) -> None:
        positions = {}
        for index, row in enumerate(self.sheet.iter_rows(self.COL_B, values_only=True), start=1):
            # skip excluded tickers
            if row[1] in self.EXCLUDED_TICKERS:
                continue

            try:
                positions[row[1]] = (index, self._repr_in_thousands(row[4]))
            except KeyError:
                logging.warning(f"Ticker: {row[1]} duplicated in exel sheet")

        self.positions = positions

    def _repr_in_thousands(self, value: str) -> float:
        """ represent value in thousands"""
        value = self._convert_to_float(value)
        try:
            return round((float(value) / 1000), 1)
        except ValueError:
            print(f"ValueError: {value} is not a valid number")

    @staticmethod
    def _convert_to_float(value: str) -> float:
        """ convert value to float None converts to 0.0"""
        if not value:
            return 0.0
        try:
            return float(value)
        except ValueError:
            logging.warning(f"ValueError: {value} is not a valid number")

    def _is_api_position_different(self, ticker: str, api_value: float) -> bool:
        file_value = self.positions[ticker][1]
        return api_value != file_value

    def _update_position(self, api_position: dict) -> None:
        """ update position to file"""
        ticker = api_position["symbolId"]
        if ticker not in self.positions.keys():
            logging.warning(f"Ticker: {ticker} not found in file")
        else:
            api_value = self._repr_in_thousands(api_position["convertedValue"])
            if self._is_api_position_different(ticker=ticker, api_value=api_value):
                self.sheet.cell(row=self.positions["ticker"][0], col=self.COL_E).value = api_value if api_value else None
                logging.info(f"UPDATE: {ticker} successful")

    @staticmethod
    def _get_currencies_converted_value(currencies: List[Dict[str]]) -> float:
        return round(sum(float(currency["convertedValue"]) for currency in currencies) / 1000, 1)

    def _update_file_exante_available_cash(self, api_currencies: List[Dict[str]]) -> None:
        api_available_cash = self._get_currencies_converted_value(api_currencies)
        file_available_cash = self._repr_in_thousands(
            self.sheet.cell(row=self.available_cash_row, column=self.COL_E).value)
        if api_available_cash != file_available_cash:
            self.sheet.cell(row=self.available_cash_row,
                            column=self.COL_E).value = api_available_cash if api_available_cash else None
            logging.info(f"UPDATE: {self.AVAILABLE_CASH} successful")
